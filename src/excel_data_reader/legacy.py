"""Legacy Excel 97-2003 (BIFF8) adapter backed by xlrd."""

from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass
from io import StringIO
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from xlrd.biffh import XLRDError
from xlrd.compdoc import CompDocError

from excel_data_reader.diagnostics import (
    Diagnostic,
    DiagnosticCode,
    ExcelDataReaderError,
    Severity,
)


class LegacyWorkbookError(ExcelDataReaderError):
    """Raised when a compound document cannot be parsed as a legacy workbook."""


@dataclass(frozen=True)
class LegacyLoadResult:
    workbook: Workbook
    diagnostics: tuple[Diagnostic, ...]


def load_legacy_workbook(
    path: str | Path,
    *,
    checkpoint: Callable[[], None] | None = None,
) -> LegacyLoadResult:
    """Load an XLS workbook into the adapter-neutral OpenPyXL worksheet surface.

    Args:
        path: Filesystem path to an Excel 97-2003 workbook.
        checkpoint: Optional callback invoked during long-running work.
    """

    _checkpoint(checkpoint)
    try:
        source = xlrd.open_workbook(
            filename=str(path),
            formatting_info=True,
            logfile=StringIO(),
            on_demand=True,
            ragged_rows=True,
            use_mmap=False,
        )
    except (CompDocError, XLRDError) as error:
        message = str(error)
        code = (
            DiagnosticCode.ENCRYPTED_WORKBOOK
            if "encrypt" in message.casefold() or "password" in message.casefold()
            else DiagnosticCode.INVALID_LEGACY_WORKBOOK
        )
        raise LegacyWorkbookError(Diagnostic(code, _legacy_error_message(code))) from error

    target = Workbook()
    try:
        if source.nsheets < 1:
            raise LegacyWorkbookError(
                Diagnostic(
                    DiagnosticCode.INVALID_LEGACY_WORKBOOK,
                    "legacy workbook does not contain a worksheet",
                )
            )
        for sheet_index in range(source.nsheets):
            _checkpoint(checkpoint)
            source_sheet = source.sheet_by_index(sheet_index)
            target_sheet = (
                target.active if sheet_index == 0 else target.create_sheet(source_sheet.name)
            )
            target_sheet.title = source_sheet.name
            target_sheet.sheet_state = _sheet_state(source_sheet.visibility)
            _copy_sheet(
                source,
                source_sheet,
                target_sheet,
                checkpoint=checkpoint,
            )
        _copy_defined_names(source, target, checkpoint=checkpoint)
    except BaseException:
        target.close()
        raise
    finally:
        source.release_resources()

    return LegacyLoadResult(
        workbook=target,
        diagnostics=(
            Diagnostic(
                DiagnosticCode.LEGACY_XLS_LIMITED,
                "legacy XLS formulas are exposed as stored results; macros, embedded objects, "
                "native tables, filters, and other unsupported BIFF features are omitted",
                severity=Severity.WARNING,
            ),
        ),
    )


def _copy_sheet(
    source_book: xlrd.book.Book,
    source_sheet: xlrd.sheet.Sheet,
    target_sheet,
    *,
    checkpoint: Callable[[], None] | None,
) -> None:
    """Copy supported BIFF worksheet values and structural metadata.

    Args:
        source_book: Source xlrd workbook used for formats and date conversion.
        source_sheet: Source xlrd worksheet to copy.
        target_sheet: Destination OpenPyXL worksheet.
        checkpoint: Optional callback invoked while rows are copied.
    """

    for row_index in range(source_sheet.nrows):
        _checkpoint(checkpoint)
        for column_index in range(source_sheet.row_len(row_index)):
            source_cell = source_sheet.cell(row_index, column_index)
            if source_cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                continue
            target_cell = target_sheet.cell(
                row=row_index + 1,
                column=column_index + 1,
                value=_cell_value(source_book, source_cell),
            )
            number_format = _number_format(source_book, source_cell.xf_index)
            if number_format:
                target_cell.number_format = number_format

    for row_index, row_info in source_sheet.rowinfo_map.items():
        if row_info.hidden:
            target_sheet.row_dimensions[row_index + 1].hidden = True
    for column_index, column_info in source_sheet.colinfo_map.items():
        if column_info.hidden:
            target_sheet.column_dimensions[get_column_letter(column_index + 1)].hidden = True
    for row_low, row_high, column_low, column_high in source_sheet.merged_cells:
        target_sheet.merge_cells(
            start_row=row_low + 1,
            end_row=row_high,
            start_column=column_low + 1,
            end_column=column_high,
        )


def _copy_defined_names(
    source: xlrd.book.Book,
    target: Workbook,
    *,
    checkpoint: Callable[[], None] | None,
) -> None:
    """Copy resolvable rectangular BIFF names into OpenPyXL definitions.

    Args:
        source: Source xlrd workbook containing BIFF names.
        target: Destination OpenPyXL workbook.
        checkpoint: Optional callback invoked while names are copied.
    """

    for source_name in source.name_obj_list:
        _checkpoint(checkpoint)
        if source_name.macro or source_name.scope < -1:
            continue
        try:
            sheet, row_low, row_high, column_low, column_high = source_name.area2d(clipped=False)
        except XLRDError:
            continue
        if row_high <= row_low or column_high <= column_low:
            continue
        start = f"${get_column_letter(column_low + 1)}${row_low + 1}"
        end = f"${get_column_letter(column_high)}${row_high}"
        definition = DefinedName(
            str(source_name.name),
            attr_text=f"{quote_sheetname(sheet.name)}!{start}:{end}",
        )
        if source_name.scope == -1:
            target.defined_names.add(definition)
        else:
            target.worksheets[source_name.scope].defined_names.add(definition)


def _cell_value(source: xlrd.book.Book, cell: xlrd.sheet.Cell):
    """Convert an xlrd cell into the corresponding public Python scalar.

    Args:
        source: Source workbook providing the date-system setting.
        cell: Source xlrd cell to convert.
    """

    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, source.datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return xlrd.error_text_from_code.get(int(cell.value), "#VALUE!")
    return cell.value


def _number_format(source: xlrd.book.Book, xf_index: int) -> str | None:
    """Return an XLS number format string when the XF record resolves.

    Args:
        source: Source workbook containing XF and format records.
        xf_index: Zero-based XF record index for the source cell.
    """

    try:
        format_key = source.xf_list[xf_index].format_key
        return str(source.format_map[format_key].format_str)
    except (IndexError, KeyError):
        return None


def _sheet_state(visibility: int) -> str:
    """Map BIFF visibility flags to OpenPyXL worksheet state values.

    Args:
        visibility: BIFF visibility flag from the source worksheet.
    """

    return {0: "visible", 1: "hidden", 2: "veryHidden"}.get(visibility, "visible")


def _legacy_error_message(code: DiagnosticCode) -> str:
    """Return a path-safe public message for a legacy parsing failure.

    Args:
        code: Stable diagnostic code identifying the parsing failure.
    """

    if code is DiagnosticCode.ENCRYPTED_WORKBOOK:
        return "password-protected legacy XLS workbooks are not supported"
    return "compound document is not a readable Excel 97-2003 workbook"


def _checkpoint(callback: Callable[[], None] | None) -> None:
    """Invoke a cooperative execution checkpoint when one is configured.

    Args:
        callback: Optional zero-argument checkpoint callback.
    """

    if callback is not None:
        callback()
