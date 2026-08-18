"""OpenPyXL adapter and public workbook reader."""

from __future__ import annotations

from collections.abc import Callable, Iterator, Sequence
from dataclasses import dataclass, replace
from itertools import product
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, range_boundaries
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from excel_data_reader.diagnostics import (
    Diagnostic,
    DiagnosticCode,
    ExcelDataReaderError,
    Severity,
)
from excel_data_reader.legacy import load_legacy_workbook
from excel_data_reader.model import (
    BodyPolicy,
    BodyPolicyMode,
    CandidateReason,
    CellData,
    ColumnInfo,
    Confidence,
    Coordinate,
    DataRow,
    DiscoveryCandidate,
    DiscoveryReport,
    FormulaValue,
    HeaderEvidence,
    MatchSet,
    MatchSource,
    NamedRangeInfo,
    NativeTableInfo,
    RangeReference,
    Rectangle,
    SheetData,
    SheetInfo,
    SheetScan,
    TableData,
    TableMatch,
    TableQuery,
    ValueMode,
    WorkbookFormat,
    WorkbookInventory,
)
from excel_data_reader.normalization import normalize_header


@dataclass(frozen=True)
class _HeaderField:
    requested: str
    normalized: str
    accepted: frozenset[str]
    required: bool


@dataclass
class _CandidateDraft:
    sheet: str
    source: MatchSource
    header_row: int | None
    bounds: Rectangle | None
    evidence: tuple[HeaderEvidence, ...]
    produced_matches: int
    reasons: list[CandidateReason]
    name: str | None = None


@dataclass
class _DiscoveryTrace:
    scans: list[SheetScan]
    candidates: list[_CandidateDraft]
    truncated_sheets: set[str]

    @classmethod
    def create(cls) -> _DiscoveryTrace:
        """Return an empty mutable trace collector for one discovery run."""

        return cls([], [], set())


class ExcelReader:
    """Discover and extract data from one workbook.

    Use :meth:`open` and a context manager so both formula and cached-value
    workbook handles are closed deterministically.
    """

    def __init__(
        self,
        path: Path,
        workbook: Any,
        cached_workbook: Any | None,
        *,
        value_mode: ValueMode,
        max_scan_cells: int,
        max_candidates: int,
        checkpoint: Callable[[], None] | None,
        workbook_format: WorkbookFormat,
        diagnostics: tuple[Diagnostic, ...],
    ) -> None:
        """Initialize a reader around already-open adapter workbooks."""

        self.path = path
        self._workbook = workbook
        self._cached_workbook = cached_workbook
        self.value_mode = value_mode
        self.max_scan_cells = max_scan_cells
        self.max_candidates = max_candidates
        self._checkpoint_callback = checkpoint
        self.workbook_format = workbook_format
        self.diagnostics = diagnostics
        self._closed = False

    @classmethod
    def open(
        cls,
        path: str | Path,
        *,
        value_mode: ValueMode | str = ValueMode.FORMULA,
        max_scan_cells: int = 2_000_000,
        max_candidates: int = 100,
        checkpoint: Callable[[], None] | None = None,
    ) -> ExcelReader:
        """Open a workbook for deterministic discovery and extraction."""

        mode = ValueMode(value_mode)
        if max_scan_cells < 1:
            raise ValueError("max_scan_cells must be positive")
        if max_candidates < 1:
            raise ValueError("max_candidates must be positive")

        if checkpoint is not None:
            checkpoint()
        workbook_path = Path(path)
        if workbook_path.suffix.casefold() == ".xls":
            legacy = load_legacy_workbook(workbook_path, checkpoint=checkpoint)
            return cls(
                workbook_path,
                legacy.workbook,
                None,
                value_mode=mode,
                max_scan_cells=max_scan_cells,
                max_candidates=max_candidates,
                checkpoint=checkpoint,
                workbook_format=WorkbookFormat.LEGACY_XLS,
                diagnostics=legacy.diagnostics,
            )
        workbook = load_workbook(
            workbook_path,
            read_only=False,
            data_only=False,
            keep_links=False,
        )
        cached_workbook = None
        try:
            if checkpoint is not None:
                checkpoint()
            if mode in {ValueMode.CACHED, ValueMode.BOTH}:
                cached_workbook = load_workbook(
                    workbook_path,
                    read_only=False,
                    data_only=True,
                    keep_links=False,
                )
                if checkpoint is not None:
                    checkpoint()
        except BaseException:
            workbook.close()
            if cached_workbook is not None:
                cached_workbook.close()
            raise
        return cls(
            workbook_path,
            workbook,
            cached_workbook,
            value_mode=mode,
            max_scan_cells=max_scan_cells,
            max_candidates=max_candidates,
            checkpoint=checkpoint,
            workbook_format=WorkbookFormat.OOXML,
            diagnostics=(),
        )

    def __enter__(self) -> ExcelReader:
        """Return this open reader for context-managed use."""

        self._require_open()
        return self

    def __exit__(self, *_: object) -> None:
        """Close all workbook handles when leaving a context manager."""

        self.close()

    def close(self) -> None:
        """Close formula and cached-value workbook handles idempotently."""

        if self._closed:
            return
        self._workbook.close()
        if self._cached_workbook is not None:
            self._cached_workbook.close()
        self._closed = True

    @property
    def sheet_names(self) -> tuple[str, ...]:
        """Return worksheet names in workbook order."""

        self._require_open()
        return tuple(self._workbook.sheetnames)

    def inventory(self) -> WorkbookInventory:
        """Return workbook-authored structure without extracting table bodies."""

        self._require_open()
        sheets: list[SheetInfo] = []
        native_tables: list[NativeTableInfo] = []
        for sheet in self._workbook.worksheets:
            self._checkpoint()
            tables = tuple(self._table_info(sheet, table) for table in sheet.tables.values())
            native_tables.extend(tables)
            sheets.append(
                SheetInfo(
                    name=sheet.title,
                    state=sheet.sheet_state,
                    apparent_bounds=self._apparent_bounds(sheet),
                    dimension=sheet.calculate_dimension(),
                    table_names=tuple(table.name for table in tables),
                    auto_filter_ref=sheet.auto_filter.ref,
                    merged_ranges=tuple(str(item) for item in sheet.merged_cells.ranges),
                )
            )
        return WorkbookInventory(
            sheets=tuple(sheets),
            native_tables=tuple(native_tables),
            named_ranges=tuple(self._iter_named_range_info()),
        )

    def read_range(
        self,
        sheet: str,
        cell_range: str,
        *,
        header: int | None = 0,
    ) -> TableData:
        """Extract one explicit rectangular A1 range."""

        worksheet = self._sheet(sheet)
        bounds = self._parse_rectangle(cell_range, sheet=sheet)
        match = self._match_from_rectangle(
            worksheet,
            bounds,
            header=header,
            source=MatchSource.EXPLICIT_RANGE,
        )
        return self.extract(match)

    def find_native_tables(
        self,
        name: str | None = None,
        *,
        sheet: str | None = None,
    ) -> MatchSet:
        """Find native Excel Tables by optional name and worksheet."""

        self._require_open()
        if sheet is not None:
            self._sheet(sheet)
        matches = tuple(
            self._table_match(worksheet, table)
            for worksheet in self._workbook.worksheets
            if sheet is None or worksheet.title == sheet
            for table in worksheet.tables.values()
            if name is None or self._table_name(table).casefold() == name.casefold()
        )
        diagnostics: tuple[Diagnostic, ...] = ()
        if not matches and name is not None:
            diagnostics = (
                Diagnostic(
                    DiagnosticCode.NATIVE_TABLE_NOT_FOUND,
                    f"native Excel Table {name!r} was not found",
                    sheet=sheet,
                ),
            )
        return MatchSet(matches, diagnostics)

    def get_table(self, name: str, *, sheet: str | None = None) -> TableData:
        """Extract exactly one native Excel Table."""

        return self.extract(self.find_native_tables(name, sheet=sheet).require_one())

    def find_named_ranges(
        self,
        name: str | None = None,
        *,
        sheet: str | None = None,
        header: int | None = 0,
    ) -> MatchSet:
        """Find rectangular destinations of defined names."""

        self._require_open()
        if sheet is not None:
            self._sheet(sheet)

        matching_infos = [
            item
            for item in self._iter_named_range_info()
            if name is None or item.name.casefold() == name.casefold()
        ]
        matches: list[TableMatch] = []
        diagnostics: list[Diagnostic] = []
        for item in matching_infos:
            self._checkpoint()
            if not item.is_resolvable:
                diagnostics.append(
                    Diagnostic(
                        (
                            DiagnosticCode.DYNAMIC_NAMED_RANGE_UNRESOLVED
                            if item.is_dynamic
                            else DiagnosticCode.NON_RECTANGULAR_NAMED_RANGE
                        ),
                        f"defined name {item.name!r} does not resolve to a rectangular range",
                        sheet=item.scope,
                    )
                )
                continue
            for destination in item.destinations:
                if sheet is not None and destination.sheet != sheet:
                    continue
                worksheet = self._sheet(destination.sheet)
                matches.append(
                    self._match_from_rectangle(
                        worksheet,
                        destination.bounds,
                        header=header,
                        source=MatchSource.NAMED_RANGE,
                        name=item.name,
                    )
                )

        if not matches and name is not None and not matching_infos:
            diagnostics.append(
                Diagnostic(
                    DiagnosticCode.NAMED_RANGE_NOT_FOUND,
                    f"defined name {name!r} was not found",
                    sheet=sheet,
                )
            )
        elif not matches and name is not None and not diagnostics:
            diagnostics.append(
                Diagnostic(
                    DiagnosticCode.NAMED_RANGE_NOT_FOUND,
                    f"defined name {name!r} has no destination on the requested worksheet",
                    sheet=sheet,
                )
            )
        return MatchSet(tuple(matches), tuple(diagnostics))

    def get_named_range(
        self,
        name: str,
        *,
        sheet: str | None = None,
        header: int | None = 0,
    ) -> TableData:
        """Extract exactly one rectangular defined-name destination."""

        match = self.find_named_ranges(name, sheet=sheet, header=header).require_one()
        return self.extract(match)

    def find_tables(
        self,
        headers: Sequence[str] | TableQuery,
        *,
        sheet: str | None = None,
        allow_non_adjacent_columns: bool | None = None,
        max_blank_rows: int | None = None,
    ) -> MatchSet:
        """Find tables from a header list or a structured :class:`TableQuery`.

        The sequence form is the original convenience API. Passing a
        ``TableQuery`` enables aliases, optional headers, scoped searches,
        disambiguation, and explicit body-boundary policies.
        """

        self._require_open()
        if isinstance(headers, TableQuery):
            if (
                sheet is not None
                or allow_non_adjacent_columns is not None
                or max_blank_rows is not None
            ):
                raise ExcelDataReaderError(
                    Diagnostic(
                        DiagnosticCode.INVALID_HEADER_QUERY,
                        "legacy keyword options cannot be combined with a TableQuery",
                    )
                )
            return self.query_tables(headers)

        adjacent = True if allow_non_adjacent_columns is None else allow_non_adjacent_columns
        blank_rows = 2 if max_blank_rows is None else max_blank_rows
        try:
            query = TableQuery(
                required_headers=(headers,) if isinstance(headers, str) else tuple(headers),
                sheet=sheet,
                allow_non_adjacent_columns=adjacent,
                body=BodyPolicy.until_blank_rows(blank_rows),
            )
        except (TypeError, ValueError) as error:
            raise ExcelDataReaderError(
                Diagnostic(
                    DiagnosticCode.INVALID_HEADER_QUERY,
                    str(error),
                )
            ) from error
        return self.query_tables(query)

    def query_tables(self, query: TableQuery) -> MatchSet:
        """Find tables satisfying a reusable structured query."""

        return self._query_tables(query)

    def explain(self, query: TableQuery) -> DiscoveryReport:
        """Run a query and report how every interesting candidate was handled."""

        trace = _DiscoveryTrace.create()
        result = self._query_tables(query, trace=trace)
        near = self._coerce_query_coordinate(query.near, sheet=query.sheet)
        return self._finalize_discovery_report(query, result, trace, near)

    def _query_tables(
        self,
        query: TableQuery,
        *,
        trace: _DiscoveryTrace | None = None,
    ) -> MatchSet:
        """Shared discovery implementation with optional trace collection."""

        self._require_open()
        fields = self._compile_table_query(query)
        within = self._coerce_query_rectangle(query.within, sheet=query.sheet)
        near = self._coerce_query_coordinate(query.near, sheet=query.sheet)
        if (
            query.body.mode is BodyPolicyMode.EXPLICIT
            and within is not None
            and (
                query.body.bottom_row is None
                or not within.top <= query.body.bottom_row <= within.bottom
            )
        ):
            raise ExcelDataReaderError(
                Diagnostic(
                    DiagnosticCode.INVALID_HEADER_QUERY,
                    "explicit bottom_row must lie inside the query's within range",
                    sheet=query.sheet,
                    address=within.a1,
                )
            )

        worksheets = (
            (self._sheet(query.sheet),)
            if query.sheet is not None
            else tuple(self._workbook.worksheets)
        )
        matches: list[TableMatch] = []
        diagnostics: list[Diagnostic] = []
        structural_signatures: set[tuple[str, int, tuple[int, ...]]] = set()

        for worksheet in worksheets:
            self._checkpoint()
            for table in worksheet.tables.values():
                self._checkpoint()
                native = self._table_match(worksheet, table)
                evidence = self._column_evidence(fields, native.columns)
                reasons: list[CandidateReason] = []
                if within is not None and not within.contains_rectangle(native.bounds):
                    reasons.append(CandidateReason.OUTSIDE_WITHIN)
                    self._record_trace_candidate(
                        trace,
                        _CandidateDraft(
                            sheet=worksheet.title,
                            source=MatchSource.NATIVE_TABLE,
                            header_row=native.header_row,
                            bounds=native.bounds,
                            evidence=evidence,
                            produced_matches=0,
                            reasons=reasons,
                            name=native.name,
                        ),
                    )
                    continue
                projected = self._project_match(
                    native,
                    fields,
                    allow_non_adjacent_columns=query.allow_non_adjacent_columns,
                )
                if self._missing_required_evidence(evidence):
                    reasons.append(CandidateReason.MISSING_REQUIRED_HEADERS)
                elif not projected:
                    reasons.append(CandidateReason.NON_ADJACENT_COLUMNS)
                self._record_trace_candidate(
                    trace,
                    _CandidateDraft(
                        sheet=worksheet.title,
                        source=MatchSource.NATIVE_TABLE,
                        header_row=native.header_row,
                        bounds=native.bounds,
                        evidence=evidence,
                        produced_matches=len(projected),
                        reasons=reasons,
                        name=native.name,
                    ),
                )
                for match in projected:
                    matches.append(match)
                    structural_signatures.add(self._header_signature(match))

        for worksheet in worksheets:
            self._checkpoint()
            scan_bounds = self._query_scan_bounds(worksheet, within)
            if scan_bounds is None:
                if trace is not None:
                    trace.scans.append(SheetScan(worksheet.title, None, 0, True))
                continue
            try:
                self._enforce_scan_limit(worksheet, bounds=scan_bounds)
            except ExcelDataReaderError as error:
                diagnostics.extend(error.diagnostics)
                if trace is not None:
                    trace.scans.append(
                        SheetScan(
                            worksheet.title,
                            scan_bounds,
                            scan_bounds.area,
                            False,
                            error.diagnostics,
                        )
                    )
                continue
            stop_sheet = False
            last_scanned_row = scan_bounds.top - 1
            for row in worksheet.iter_rows(
                min_row=scan_bounds.top,
                max_row=scan_bounds.bottom,
                min_col=scan_bounds.left,
                max_col=scan_bounds.right,
            ):
                self._checkpoint()
                header_row = row[0].row
                last_scanned_row = header_row
                positions: list[list[tuple[int, Any]]] = [[] for _ in fields]
                for cell in row:
                    if cell.value is None:
                        continue
                    canonical = normalize_header(cell.value)
                    for index, field in enumerate(fields):
                        if canonical in field.accepted:
                            positions[index].append((cell.column, cell.value))
                if not any(positions):
                    continue

                evidence = self._position_evidence(fields, positions, header_row)
                evidence_bounds = self._evidence_bounds(evidence, header_row)
                if self._missing_required_evidence(evidence):
                    self._record_trace_candidate(
                        trace,
                        _CandidateDraft(
                            sheet=worksheet.title,
                            source=MatchSource.HEADER,
                            header_row=header_row,
                            bounds=evidence_bounds,
                            evidence=evidence,
                            produced_matches=0,
                            reasons=[CandidateReason.MISSING_REQUIRED_HEADERS],
                        ),
                    )
                    continue

                active = tuple(
                    (field, positions[index])
                    for index, field in enumerate(fields)
                    if field.required or positions[index]
                )
                duplicate = any(len(found) > 1 for _, found in active)
                match_diagnostics: tuple[Diagnostic, ...] = ()
                if duplicate:
                    warning = Diagnostic(
                        DiagnosticCode.DUPLICATE_HEADER,
                        "a requested or optional header appears more than once on this row",
                        severity=Severity.WARNING,
                        sheet=worksheet.title,
                        address=Coordinate(row[0].row, scan_bounds.left).a1,
                    )
                    diagnostics.append(warning)
                    match_diagnostics = (warning,)

                produced_matches = 0
                rejected_non_adjacent = False
                rejected_explicit_bottom = False
                shadowed_by_native = False
                candidate_bottom = header_row
                for selected in product(*(found for _, found in active)):
                    selected_columns = tuple(item[0] for item in selected)
                    if len(set(selected_columns)) != len(selected_columns):
                        continue
                    if not query.allow_non_adjacent_columns and not self._columns_are_adjacent(
                        selected_columns
                    ):
                        rejected_non_adjacent = True
                        continue
                    columns = tuple(
                        ColumnInfo(
                            name=self._column_name(raw, index),
                            source_column=column,
                            raw_header=raw,
                            requested_header=field.requested,
                            header_coordinate=Coordinate(header_row, column),
                        )
                        for index, ((field, _), (column, raw)) in enumerate(
                            zip(active, selected, strict=True), start=1
                        )
                    )
                    if (
                        query.body.mode is BodyPolicyMode.EXPLICIT
                        and query.body.bottom_row is not None
                        and query.body.bottom_row < header_row
                    ):
                        rejected_explicit_bottom = True
                        continue
                    bottom = self._infer_body_bottom(
                        worksheet,
                        header_row,
                        selected_columns,
                        policy=query.body,
                        max_row=scan_bounds.bottom,
                    )
                    candidate_bottom = max(candidate_bottom, bottom)
                    produced_matches += 1
                    match = TableMatch(
                        sheet=worksheet.title,
                        bounds=Rectangle(
                            header_row,
                            min(selected_columns),
                            max(header_row, bottom),
                            max(selected_columns),
                        ),
                        columns=columns,
                        data_start_row=header_row + 1,
                        data_end_row=bottom,
                        source=MatchSource.HEADER,
                        confidence=Confidence.HIGH,
                        header_row=header_row,
                        diagnostics=match_diagnostics,
                    )
                    if self._header_signature(match) in structural_signatures:
                        shadowed_by_native = True
                        continue
                    matches.append(match)
                    if len(matches) >= self.max_candidates:
                        diagnostics.append(
                            Diagnostic(
                                DiagnosticCode.SCAN_LIMIT_EXCEEDED,
                                f"candidate limit of {self.max_candidates} was reached",
                                sheet=worksheet.title,
                            )
                        )
                        stop_sheet = True
                        break

                reasons: list[CandidateReason] = []
                if produced_matches == 0:
                    if rejected_non_adjacent:
                        reasons.append(CandidateReason.NON_ADJACENT_COLUMNS)
                    if rejected_explicit_bottom:
                        reasons.append(CandidateReason.EXPLICIT_BOTTOM_BEFORE_HEADER)
                elif shadowed_by_native and not any(
                    match.source is MatchSource.HEADER
                    and match.sheet == worksheet.title
                    and match.header_row == header_row
                    for match in matches
                ):
                    reasons.append(CandidateReason.SHADOWED_BY_NATIVE_TABLE)
                self._record_trace_candidate(
                    trace,
                    _CandidateDraft(
                        sheet=worksheet.title,
                        source=MatchSource.HEADER,
                        header_row=header_row,
                        bounds=(
                            None
                            if evidence_bounds is None
                            else Rectangle(
                                evidence_bounds.top,
                                evidence_bounds.left,
                                candidate_bottom,
                                evidence_bounds.right,
                            )
                        ),
                        evidence=evidence,
                        produced_matches=produced_matches,
                        reasons=reasons,
                    ),
                )
                if stop_sheet:
                    break

            if trace is not None:
                completed = not stop_sheet and worksheet.title not in trace.truncated_sheets
                scanned_bounds = (
                    scan_bounds
                    if last_scanned_row >= scan_bounds.bottom
                    else Rectangle(
                        scan_bounds.top,
                        scan_bounds.left,
                        max(scan_bounds.top, last_scanned_row),
                        scan_bounds.right,
                    )
                )
                scan_diagnostics: tuple[Diagnostic, ...] = ()
                if worksheet.title in trace.truncated_sheets:
                    scan_diagnostics = (
                        Diagnostic(
                            DiagnosticCode.SCAN_LIMIT_EXCEEDED,
                            f"explanation candidate limit of {self.max_candidates} was reached",
                            severity=Severity.WARNING,
                            sheet=worksheet.title,
                        ),
                    )
                trace.scans.append(
                    SheetScan(
                        worksheet.title,
                        scanned_bounds,
                        scanned_bounds.area,
                        completed,
                        scan_diagnostics,
                    )
                )

        matches = self._deduplicate_matches(matches)
        if near is not None and matches:
            minimum = min(self._distance_to_match(near, match) for match in matches)
            matches = [
                match for match in matches if self._distance_to_match(near, match) == minimum
            ]
        if not matches:
            diagnostics.append(
                Diagnostic(
                    DiagnosticCode.TABLE_NOT_FOUND,
                    "no table contained all requested headers",
                    sheet=query.sheet,
                )
            )
        return MatchSet(tuple(matches), tuple(diagnostics))

    def extract(self, match: TableMatch) -> TableData:
        """Extract the selected logical columns from a discovered match."""

        worksheet = self._sheet(match.sheet)
        hidden_rows = self._hidden_rows(worksheet)
        hidden_columns = self._hidden_columns(worksheet)
        rows: list[DataRow] = []
        if not match.is_empty:
            for row_index in range(match.data_start_row, match.data_end_row + 1):
                self._checkpoint()
                cells = tuple(
                    self._cell_data(
                        worksheet,
                        row_index,
                        column.source_column,
                        hidden_rows=hidden_rows,
                        hidden_columns=hidden_columns,
                    )
                    for column in match.columns
                )
                rows.append(DataRow(row_index, cells))
        return TableData(match, tuple(rows))

    def read_sheet(
        self,
        sheet: str,
        *,
        include_styled_blanks: bool = False,
    ) -> SheetData:
        """Return a sparse, coordinate-preserving worksheet snapshot."""

        worksheet = self._sheet(sheet)
        self._enforce_scan_limit(worksheet)
        hidden_rows = self._hidden_rows(worksheet)
        hidden_columns = self._hidden_columns(worksheet)
        cells: list[CellData] = []
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            self._checkpoint()
            for cell in row:
                if cell.value is None and not (include_styled_blanks and cell.has_style):
                    continue
                cells.append(
                    self._cell_data(
                        worksheet,
                        cell.row,
                        cell.column,
                        hidden_rows=hidden_rows,
                        hidden_columns=hidden_columns,
                    )
                )
        if not cells:
            bounds = None
        else:
            rows = [cell.coordinate.row for cell in cells]
            columns = [cell.coordinate.column for cell in cells]
            bounds = Rectangle(min(rows), min(columns), max(rows), max(columns))
        return SheetData(worksheet.title, tuple(cells), bounds)

    def _require_open(self) -> None:
        if self._closed:
            raise ExcelDataReaderError(
                Diagnostic(DiagnosticCode.READER_CLOSED, "the workbook reader is closed")
            )

    def _checkpoint(self) -> None:
        if self._checkpoint_callback is not None:
            self._checkpoint_callback()

    def _sheet(self, name: str) -> Worksheet:
        self._require_open()
        if name not in self._workbook.sheetnames:
            raise ExcelDataReaderError(
                Diagnostic(
                    DiagnosticCode.SHEET_NOT_FOUND,
                    f"worksheet {name!r} was not found",
                    sheet=name,
                )
            )
        return self._workbook[name]

    def _cached_sheet(self, name: str) -> Worksheet | None:
        if self._cached_workbook is None:
            return None
        return self._cached_workbook[name]

    def _cell_data(
        self,
        worksheet: Worksheet,
        row: int,
        column: int,
        *,
        hidden_rows: set[int],
        hidden_columns: set[int],
    ) -> CellData:
        formula_cell = worksheet.cell(row, column)
        cached_sheet = self._cached_sheet(worksheet.title)
        cached_cell = cached_sheet.cell(row, column) if cached_sheet is not None else None
        is_formula = formula_cell.data_type == "f"
        formula = formula_cell.value if is_formula else None
        cached_value = (
            cached_cell.value
            if cached_cell is not None
            else (None if is_formula else formula_cell.value)
        )
        if self.value_mode is ValueMode.FORMULA:
            value = formula_cell.value
        elif self.value_mode is ValueMode.CACHED:
            value = cached_value
        elif is_formula:
            value = FormulaValue(formula, cached_value)
        else:
            value = formula_cell.value
        date_cell = (
            cached_cell
            if cached_cell is not None and cached_cell.value is not None
            else formula_cell
        )
        return CellData(
            sheet=worksheet.title,
            coordinate=Coordinate(row, column),
            value=value,
            formula=formula,
            cached_value=cached_value,
            data_type=str(formula_cell.data_type),
            number_format=str(formula_cell.number_format),
            is_date=bool(date_cell.is_date),
            hidden_row=row in hidden_rows,
            hidden_column=column in hidden_columns,
        )

    def _match_from_rectangle(
        self,
        worksheet: Worksheet,
        bounds: Rectangle,
        *,
        header: int | None,
        source: MatchSource,
        name: str | None = None,
    ) -> TableMatch:
        if header is not None and (header < 0 or bounds.top + header > bounds.bottom):
            raise ExcelDataReaderError(
                Diagnostic(
                    DiagnosticCode.INVALID_RANGE,
                    "header offset lies outside the selected range",
                    sheet=worksheet.title,
                    address=bounds.a1,
                )
            )
        header_row = None if header is None else bounds.top + header
        columns: list[ColumnInfo] = []
        for index, column in enumerate(range(bounds.left, bounds.right + 1), start=1):
            raw = None if header_row is None else worksheet.cell(header_row, column).value
            columns.append(
                ColumnInfo(
                    name=self._column_name(raw, index),
                    source_column=column,
                    raw_header=raw,
                    header_coordinate=(
                        None if header_row is None else Coordinate(header_row, column)
                    ),
                )
            )
        return TableMatch(
            sheet=worksheet.title,
            bounds=bounds,
            columns=tuple(columns),
            data_start_row=bounds.top if header_row is None else header_row + 1,
            data_end_row=bounds.bottom,
            source=source,
            confidence=Confidence.STRUCTURAL,
            header_row=header_row,
            name=name,
        )

    def _table_match(self, worksheet: Worksheet, table: Table) -> TableMatch:
        bounds = self._parse_rectangle(str(table.ref), sheet=worksheet.title)
        header_count = self._optional_int(table.headerRowCount)
        totals_count = self._optional_int(table.totalsRowCount)
        if totals_count == 0 and table.totalsRowShown:
            totals_count = 1
        header_row = bounds.top if header_count else None
        table_names = tuple(str(item) for item in table.column_names)
        columns: list[ColumnInfo] = []
        for index, column in enumerate(range(bounds.left, bounds.right + 1), start=1):
            raw = None if header_row is None else worksheet.cell(header_row, column).value
            if (raw is None or not normalize_header(raw)) and index <= len(table_names):
                raw = table_names[index - 1]
            columns.append(
                ColumnInfo(
                    name=self._column_name(raw, index),
                    source_column=column,
                    raw_header=raw,
                    header_coordinate=(
                        None if header_row is None else Coordinate(header_row, column)
                    ),
                )
            )
        data_start = bounds.top + header_count
        data_end = bounds.bottom - totals_count
        return TableMatch(
            sheet=worksheet.title,
            bounds=bounds,
            columns=tuple(columns),
            data_start_row=data_start,
            data_end_row=data_end,
            source=MatchSource.NATIVE_TABLE,
            confidence=Confidence.STRUCTURAL,
            header_row=header_row,
            name=self._table_name(table),
        )

    def _project_match(
        self,
        match: TableMatch,
        fields: tuple[_HeaderField, ...],
        *,
        allow_non_adjacent_columns: bool,
    ) -> tuple[TableMatch, ...]:
        positions: list[list[ColumnInfo]] = [[] for _ in fields]
        for column in match.columns:
            canonical = normalize_header(column.name)
            for index, field in enumerate(fields):
                if canonical in field.accepted:
                    positions[index].append(column)
        if any(field.required and not positions[index] for index, field in enumerate(fields)):
            return ()
        active = tuple(
            (field, positions[index])
            for index, field in enumerate(fields)
            if field.required or positions[index]
        )
        projected: list[TableMatch] = []
        for selected in product(*(found for _, found in active)):
            physical = tuple(column.source_column for column in selected)
            if len(set(physical)) != len(physical):
                continue
            if not allow_non_adjacent_columns and not self._columns_are_adjacent(physical):
                continue
            columns = tuple(
                replace(column, requested_header=field.requested)
                for (field, _), column in zip(active, selected, strict=True)
            )
            diagnostics = match.diagnostics
            if any(len(found) > 1 for _, found in active):
                diagnostics += (
                    Diagnostic(
                        DiagnosticCode.DUPLICATE_HEADER,
                        "a requested or optional header occurs more than once in the table",
                        severity=Severity.WARNING,
                        sheet=match.sheet,
                        address=match.range,
                    ),
                )
            projected.append(replace(match, columns=columns, diagnostics=diagnostics))
        return tuple(projected)

    @staticmethod
    def _column_evidence(
        fields: tuple[_HeaderField, ...],
        columns: tuple[ColumnInfo, ...],
    ) -> tuple[HeaderEvidence, ...]:
        evidence: list[HeaderEvidence] = []
        for field in fields:
            matching = tuple(
                column for column in columns if normalize_header(column.name) in field.accepted
            )
            evidence.append(
                HeaderEvidence(
                    requested_header=field.requested,
                    required=field.required,
                    coordinates=tuple(
                        column.header_coordinate
                        for column in matching
                        if column.header_coordinate is not None
                    ),
                    raw_headers=tuple(str(column.raw_header or column.name) for column in matching),
                )
            )
        return tuple(evidence)

    @staticmethod
    def _position_evidence(
        fields: tuple[_HeaderField, ...],
        positions: list[list[tuple[int, Any]]],
        header_row: int,
    ) -> tuple[HeaderEvidence, ...]:
        return tuple(
            HeaderEvidence(
                requested_header=field.requested,
                required=field.required,
                coordinates=tuple(Coordinate(header_row, column) for column, _ in positions[index]),
                raw_headers=tuple(str(raw) for _, raw in positions[index]),
            )
            for index, field in enumerate(fields)
        )

    @staticmethod
    def _missing_required_evidence(evidence: tuple[HeaderEvidence, ...]) -> bool:
        return any(item.required and not item.matched for item in evidence)

    @staticmethod
    def _evidence_bounds(
        evidence: tuple[HeaderEvidence, ...],
        header_row: int,
    ) -> Rectangle | None:
        columns = [coordinate.column for item in evidence for coordinate in item.coordinates]
        if not columns:
            return None
        return Rectangle(header_row, min(columns), header_row, max(columns))

    def _record_trace_candidate(
        self,
        trace: _DiscoveryTrace | None,
        candidate: _CandidateDraft,
    ) -> None:
        if trace is None:
            return
        if len(trace.candidates) >= self.max_candidates:
            trace.truncated_sheets.add(candidate.sheet)
            return
        trace.candidates.append(candidate)

    def _finalize_discovery_report(
        self,
        query: TableQuery,
        result: MatchSet,
        trace: _DiscoveryTrace,
        near: Coordinate | None,
    ) -> DiscoveryReport:
        candidates: list[DiscoveryCandidate] = []
        for draft in trace.candidates:
            selected = any(
                match.sheet == draft.sheet
                and match.source is draft.source
                and (
                    (draft.source is MatchSource.NATIVE_TABLE and match.name == draft.name)
                    or (draft.source is MatchSource.HEADER and match.header_row == draft.header_row)
                )
                for match in result.matches
            )
            reasons = list(draft.reasons)
            if draft.produced_matches and not selected and not reasons:
                if near is not None:
                    reasons.append(CandidateReason.FARTHER_FROM_NEAR)
                elif any(
                    match.source is MatchSource.NATIVE_TABLE
                    and match.sheet == draft.sheet
                    and match.header_row == draft.header_row
                    for match in result.matches
                ):
                    reasons.append(CandidateReason.SHADOWED_BY_NATIVE_TABLE)
                else:
                    reasons.append(CandidateReason.CANDIDATE_LIMIT)
            distance = (
                None
                if near is None or draft.bounds is None
                else self._distance_to_rectangle(near, draft.bounds)
            )
            candidates.append(
                DiscoveryCandidate(
                    sheet=draft.sheet,
                    source=draft.source,
                    header_row=draft.header_row,
                    bounds=draft.bounds,
                    evidence=draft.evidence,
                    produced_matches=draft.produced_matches,
                    selected=selected,
                    reasons=tuple(dict.fromkeys(reasons)),
                    name=draft.name,
                    distance_from_near=distance,
                )
            )
        return DiscoveryReport(query, result, tuple(trace.scans), tuple(candidates))

    def _infer_body_bottom(
        self,
        worksheet: Worksheet,
        header_row: int,
        columns: tuple[int, ...],
        *,
        policy: BodyPolicy,
        max_row: int,
    ) -> int:
        if policy.mode is BodyPolicyMode.EXPLICIT:
            if policy.bottom_row is None:
                raise AssertionError("validated explicit policy has no bottom row")
            return policy.bottom_row

        last_nonblank = header_row
        blank_run = 0
        for row in range(header_row + 1, max_row + 1):
            self._checkpoint()
            if any(worksheet.cell(row, column).value is not None for column in columns):
                last_nonblank = row
                blank_run = 0
                continue
            if policy.mode is BodyPolicyMode.LAST_POPULATED:
                continue
            blank_run += 1
            if blank_run >= policy.blank_rows:
                break
        return last_nonblank

    def _compile_table_query(self, query: TableQuery) -> tuple[_HeaderField, ...]:
        required_headers = tuple(query.required_headers)
        optional_headers = tuple(query.optional_headers)
        declared = required_headers + optional_headers
        normalized = tuple(normalize_header(header) for header in declared)
        if not required_headers or any(not header for header in normalized):
            raise ExcelDataReaderError(
                Diagnostic(
                    DiagnosticCode.INVALID_HEADER_QUERY,
                    "at least one non-empty required header is needed, and optional headers "
                    "cannot be empty",
                )
            )
        duplicates = sorted({item for item in normalized if normalized.count(item) > 1})
        if duplicates:
            raise ExcelDataReaderError(
                Diagnostic(
                    DiagnosticCode.INVALID_HEADER_QUERY,
                    "requested headers are not unique after normalization: "
                    + ", ".join(duplicates),
                )
            )

        field_by_name = {name: index for index, name in enumerate(normalized)}
        accepted = [{name} for name in normalized]
        owner = dict(field_by_name)
        for raw_key, raw_aliases in query.aliases.items():
            key = normalize_header(raw_key)
            if key not in field_by_name:
                raise ExcelDataReaderError(
                    Diagnostic(
                        DiagnosticCode.INVALID_HEADER_QUERY,
                        f"alias key {raw_key!r} is not a declared header",
                    )
                )
            field_index = field_by_name[key]
            for raw_alias in raw_aliases:
                alias = normalize_header(raw_alias)
                if not alias:
                    raise ExcelDataReaderError(
                        Diagnostic(
                            DiagnosticCode.INVALID_HEADER_QUERY,
                            f"alias for {raw_key!r} cannot be empty",
                        )
                    )
                previous = owner.get(alias)
                if previous is not None and previous != field_index:
                    raise ExcelDataReaderError(
                        Diagnostic(
                            DiagnosticCode.INVALID_HEADER_QUERY,
                            f"normalized alias {alias!r} belongs to more than one header",
                        )
                    )
                owner[alias] = field_index
                accepted[field_index].add(alias)

        required_count = len(required_headers)
        return tuple(
            _HeaderField(
                requested=str(requested),
                normalized=normalized[index],
                accepted=frozenset(accepted[index]),
                required=index < required_count,
            )
            for index, requested in enumerate(declared)
        )

    def _iter_named_range_info(self) -> Iterator[NamedRangeInfo]:
        for name, definition in self._workbook.defined_names.items():
            self._checkpoint()
            yield self._named_range_info(name, definition, scope=None)
        for worksheet in self._workbook.worksheets:
            self._checkpoint()
            for name, definition in worksheet.defined_names.items():
                self._checkpoint()
                yield self._named_range_info(name, definition, scope=worksheet.title)

    def _named_range_info(
        self,
        name: str,
        definition: Any,
        *,
        scope: str | None,
    ) -> NamedRangeInfo:
        destinations: list[RangeReference] = []
        unresolved = False
        try:
            raw_destinations = tuple(definition.destinations)
        except (AttributeError, TypeError, ValueError):
            raw_destinations = ()
            unresolved = True
        for sheet, cell_range in raw_destinations:
            try:
                bounds = self._parse_rectangle(str(cell_range), sheet=str(sheet))
            except ExcelDataReaderError:
                unresolved = True
                continue
            destinations.append(RangeReference(str(sheet), bounds))
        if not destinations:
            unresolved = True
        return NamedRangeInfo(
            name=str(name),
            scope=scope,
            value=str(definition.value),
            destinations=tuple(destinations),
            is_dynamic=str(getattr(definition, "type", "")).upper() == "FUNC",
            is_resolvable=bool(destinations) and not unresolved,
        )

    def _table_info(self, worksheet: Worksheet, table: Table) -> NativeTableInfo:
        bounds = self._parse_rectangle(str(table.ref), sheet=worksheet.title)
        totals_count = self._optional_int(table.totalsRowCount)
        if totals_count == 0 and table.totalsRowShown:
            totals_count = 1
        return NativeTableInfo(
            name=self._table_name(table),
            sheet=worksheet.title,
            bounds=bounds,
            column_names=tuple(str(item) for item in table.column_names),
            header_row_count=self._optional_int(table.headerRowCount),
            totals_row_count=totals_count,
        )

    @staticmethod
    def _table_name(table: Table) -> str:
        return str(table.displayName or table.name)

    @staticmethod
    def _optional_int(value: Any) -> int:
        return 0 if value is None else int(value)

    @staticmethod
    def _column_name(raw: Any | None, position: int) -> str:
        if raw is None or not normalize_header(raw):
            return f"column_{position}"
        return str(raw)

    @staticmethod
    def _columns_are_adjacent(columns: tuple[int, ...]) -> bool:
        ordered = sorted(columns)
        return ordered[-1] - ordered[0] + 1 == len(ordered)

    @staticmethod
    def _header_signature(match: TableMatch) -> tuple[str, int, tuple[int, ...]]:
        return (
            match.sheet,
            match.header_row or match.bounds.top,
            tuple(column.source_column for column in match.columns),
        )

    @staticmethod
    def _deduplicate_matches(matches: list[TableMatch]) -> list[TableMatch]:
        unique: dict[tuple[str, int, tuple[int, ...]], TableMatch] = {}
        for match in matches:
            signature = ExcelReader._header_signature(match)
            existing = unique.get(signature)
            if existing is None or match.source is MatchSource.NATIVE_TABLE:
                unique[signature] = match
        return list(unique.values())

    def _parse_rectangle(self, value: str, *, sheet: str) -> Rectangle:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(value)
            if None in {min_col, min_row, max_col, max_row}:
                raise ValueError("range must have finite row and column bounds")
            return Rectangle(int(min_row), int(min_col), int(max_row), int(max_col))
        except (TypeError, ValueError) as error:
            raise ExcelDataReaderError(
                Diagnostic(
                    DiagnosticCode.INVALID_RANGE,
                    f"{value!r} is not one finite rectangular A1 range",
                    sheet=sheet,
                )
            ) from error

    def _coerce_query_rectangle(
        self,
        value: Rectangle | str | None,
        *,
        sheet: str | None,
    ) -> Rectangle | None:
        if value is None or isinstance(value, Rectangle):
            return value
        if not isinstance(value, str):
            raise ExcelDataReaderError(
                Diagnostic(
                    DiagnosticCode.INVALID_RANGE,
                    "within must be a Rectangle or finite rectangular A1 range",
                    sheet=sheet,
                )
            )
        return self._parse_rectangle(value, sheet=sheet or "<query>")

    def _coerce_query_coordinate(
        self,
        value: Coordinate | str | None,
        *,
        sheet: str | None,
    ) -> Coordinate | None:
        if value is None or isinstance(value, Coordinate):
            return value
        if not isinstance(value, str):
            raise ExcelDataReaderError(
                Diagnostic(
                    DiagnosticCode.INVALID_RANGE,
                    "near must be a Coordinate or one A1 cell address",
                    sheet=sheet,
                )
            )
        bounds = self._parse_rectangle(value, sheet=sheet or "<query>")
        if bounds.area != 1:
            raise ExcelDataReaderError(
                Diagnostic(
                    DiagnosticCode.INVALID_RANGE,
                    "near must identify exactly one cell",
                    sheet=sheet,
                    address=bounds.a1,
                )
            )
        return Coordinate(bounds.top, bounds.left)

    @staticmethod
    def _query_scan_bounds(
        worksheet: Worksheet,
        within: Rectangle | None,
    ) -> Rectangle | None:
        apparent = Rectangle(1, 1, int(worksheet.max_row), int(worksheet.max_column))
        if within is None:
            return apparent
        top = max(apparent.top, within.top)
        left = max(apparent.left, within.left)
        bottom = min(apparent.bottom, within.bottom)
        right = min(apparent.right, within.right)
        if bottom < top or right < left:
            return None
        return Rectangle(top, left, bottom, right)

    @staticmethod
    def _distance_to_match(coordinate: Coordinate, match: TableMatch) -> int:
        return ExcelReader._distance_to_rectangle(coordinate, match.bounds)

    @staticmethod
    def _distance_to_rectangle(coordinate: Coordinate, bounds: Rectangle) -> int:
        row_distance = max(
            bounds.top - coordinate.row,
            0,
            coordinate.row - bounds.bottom,
        )
        column_distance = max(
            bounds.left - coordinate.column,
            0,
            coordinate.column - bounds.right,
        )
        return row_distance + column_distance

    def _enforce_scan_limit(
        self,
        worksheet: Worksheet,
        *,
        bounds: Rectangle | None = None,
    ) -> None:
        scan_bounds = bounds or Rectangle(
            1,
            1,
            int(worksheet.max_row),
            int(worksheet.max_column),
        )
        apparent_cells = scan_bounds.area
        if apparent_cells > self.max_scan_cells:
            raise ExcelDataReaderError(
                Diagnostic(
                    DiagnosticCode.SCAN_LIMIT_EXCEEDED,
                    f"scan area is {apparent_cells:,} cells; limit is {self.max_scan_cells:,}",
                    sheet=worksheet.title,
                    address=scan_bounds.a1,
                )
            )

    @staticmethod
    def _apparent_bounds(worksheet: Worksheet) -> Rectangle | None:
        if (
            worksheet.max_row == 1
            and worksheet.max_column == 1
            and worksheet["A1"].value is None
            and not worksheet["A1"].has_style
            and not worksheet.merged_cells.ranges
        ):
            return None
        return Rectangle(1, 1, int(worksheet.max_row), int(worksheet.max_column))

    @staticmethod
    def _hidden_rows(worksheet: Worksheet) -> set[int]:
        return {
            int(index) for index, dimension in worksheet.row_dimensions.items() if dimension.hidden
        }

    @staticmethod
    def _hidden_columns(worksheet: Worksheet) -> set[int]:
        hidden: set[int] = set()
        for key, dimension in worksheet.column_dimensions.items():
            if not dimension.hidden:
                continue
            minimum = int(dimension.min or column_index_from_string(key))
            maximum = int(dimension.max or minimum)
            hidden.update(range(minimum, maximum + 1))
        return hidden
