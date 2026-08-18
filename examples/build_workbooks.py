"""Build the checked-in example workbooks with OpenPyXL."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from xlwt import Workbook as LegacyWorkbook
from xlwt import easyxf

WORKBOOK_DIR = Path(__file__).parent / "workbooks"

NAVY = "1F4E78"
TEAL = "0F766E"
SLATE = "475569"
LIGHT_BLUE = "D9EAF7"
LIGHT_TEAL = "CCFBF1"
LIGHT_GRAY = "E2E8F0"
WHITE = "FFFFFF"
TEXT = "1E293B"
THIN_GRAY = Side(style="thin", color="CBD5E1")


def _configure_sheet(sheet: Worksheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_view.zoomScale = 95


def _add_title(sheet: Worksheet, title: str, subtitle: str, *, last_column: int) -> None:
    last = get_column_letter(last_column)
    sheet.merge_cells(f"A1:{last}1")
    sheet.merge_cells(f"A2:{last}2")
    title_cell = sheet["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    title_cell.fill = PatternFill(fill_type="solid", fgColor=NAVY)
    title_cell.alignment = Alignment(vertical="center")
    subtitle_cell = sheet["A2"]
    subtitle_cell.value = subtitle
    subtitle_cell.font = Font(name="Aptos", size=10, color=TEXT)
    subtitle_cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_BLUE)
    subtitle_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 28
    sheet.row_dimensions[3].height = 8


def _style_header(sheet: Worksheet, row: int, columns: list[int], *, color: str = TEAL) -> None:
    for column in columns:
        cell = sheet.cell(row, column)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill(fill_type="solid", fgColor=color)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=color))
    sheet.row_dimensions[row].height = 28


def _style_body(sheet: Worksheet, cell_range: str) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=THIN_GRAY)


def _set_widths(sheet: Worksheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def build_native_table(output_dir: Path) -> Path:
    """Create a workbook containing an authored native Excel Table."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    _configure_sheet(sheet)
    _add_title(
        sheet,
        "Orders — Native Excel Table",
        "Use ExcelReader.get_table('OrdersTable'); the totals row is metadata, not a data row.",
        last_column=6,
    )

    sheet.append([])
    sheet.append([])
    headers = ["Order ID", "Customer", "Invoice Date", "Units", "Unit Price", "Amount"]
    for column, value in enumerate(headers, start=1):
        sheet.cell(4, column, value)
    rows = [
        ("ORD-1001", "Acme Labs", date(2026, 1, 5), 3, 125.0),
        ("ORD-1002", "Northwind", date(2026, 1, 12), 2, 210.0),
        ("ORD-1003", "Globex", date(2026, 2, 3), 5, 89.5),
        ("ORD-1004", "Initech", date(2026, 2, 18), 1, 760.0),
    ]
    for row_index, row in enumerate(rows, start=5):
        for column, value in enumerate(row, start=1):
            sheet.cell(row_index, column, value)
        sheet.cell(row_index, 6, f"=D{row_index}*E{row_index}")

    sheet["A9"] = "Total"
    sheet["F9"] = "=SUBTOTAL(109,F5:F8)"
    _style_header(sheet, 4, list(range(1, 7)))
    _style_body(sheet, "A5:F9")
    for row in range(5, 9):
        sheet.cell(row, 3).number_format = "yyyy-mm-dd"
        sheet.cell(row, 4).number_format = "#,##0"
        sheet.cell(row, 5).number_format = '"$"#,##0.00'
        sheet.cell(row, 6).number_format = '"$"#,##0.00'
        sheet.row_dimensions[row].height = 22
    sheet["A9"].font = Font(name="Aptos", size=10, bold=True, color=TEXT)
    sheet["F9"].font = Font(name="Aptos", size=10, bold=True, color=TEXT)
    sheet["F9"].number_format = '"$"#,##0.00'
    sheet["A9"].fill = PatternFill(fill_type="solid", fgColor=LIGHT_BLUE)
    sheet["F9"].fill = PatternFill(fill_type="solid", fgColor=LIGHT_BLUE)
    _set_widths(
        sheet,
        {"A": 15, "B": 22, "C": 16, "D": 10, "E": 14, "F": 15},
    )

    table = Table(
        displayName="OrdersTable",
        ref="A4:F9",
        totalsRowCount=1,
        totalsRowShown=True,
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "native_table.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def build_scattered_headers(output_dir: Path) -> Path:
    """Create a workbook with requested columns separated by unrelated columns."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scattered Orders"
    _configure_sheet(sheet)
    _add_title(
        sheet,
        "Orders — Scattered Header Columns",
        "Customer ID, Invoice-Date, and Amount are separated; row 7 is an internal blank row.",
        last_column=7,
    )

    headers = ["Customer ID", "Region", None, "Invoice-Date", "Owner", None, "Amount"]
    for column, value in enumerate(headers, start=1):
        sheet.cell(4, column, value)
    data = {
        5: ("C-001", "North", date(2026, 3, 1), "Ava", 1250.0),
        6: ("C-002", "South", date(2026, 3, 4), "Ben", 840.0),
        8: ("C-003", "West", date(2026, 3, 9), "Chloe", 2190.0),
    }
    for row, values in data.items():
        customer_id, region, invoice_date, owner, amount = values
        sheet.cell(row, 1, customer_id)
        sheet.cell(row, 2, region)
        sheet.cell(row, 4, invoice_date)
        sheet.cell(row, 5, owner)
        sheet.cell(row, 7, amount)
        sheet.cell(row, 4).number_format = "yyyy-mm-dd"
        sheet.cell(row, 7).number_format = '"$"#,##0.00'
        sheet.row_dimensions[row].height = 22

    _style_header(sheet, 4, [1, 4, 7])
    _style_header(sheet, 4, [2, 5], color=SLATE)
    _style_body(sheet, "A5:G8")
    sheet.row_dimensions[7].height = 10
    _set_widths(
        sheet,
        {"A": 16, "B": 14, "C": 5, "D": 16, "E": 16, "F": 5, "G": 15},
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "scattered_headers.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def build_named_and_headerless(output_dir: Path) -> Path:
    """Create named-range and headerless-range examples in one workbook."""

    workbook = Workbook()
    inventory = workbook.active
    inventory.title = "Inventory"
    _configure_sheet(inventory)
    _add_title(
        inventory,
        "Inventory — Defined Name",
        "The rectangular defined name InventoryData refers to B4:E8.",
        last_column=5,
    )
    headers = ["SKU", "Description", "Quantity", "Unit Cost"]
    for column, value in enumerate(headers, start=2):
        inventory.cell(4, column, value)
    rows = [
        ("SKU-100", "USB-C hub", 18, 42.5),
        ("SKU-110", "Laptop stand", 9, 31.0),
        ("SKU-120", "Wireless mouse", 24, 27.75),
        ("SKU-130", "Mechanical keyboard", 7, 88.0),
    ]
    for row_index, row in enumerate(rows, start=5):
        for column, value in enumerate(row, start=2):
            inventory.cell(row_index, column, value)
        inventory.cell(row_index, 4).number_format = "#,##0"
        inventory.cell(row_index, 5).number_format = '"$"#,##0.00'
        inventory.row_dimensions[row_index].height = 22
    _style_header(inventory, 4, list(range(2, 6)))
    _style_body(inventory, "B5:E8")
    _set_widths(inventory, {"A": 4, "B": 16, "C": 25, "D": 12, "E": 14})

    raw = workbook.create_sheet("Raw Import")
    _configure_sheet(raw)
    _add_title(
        raw,
        "Raw Import — No Header Row",
        "Read C5:F8 with header=None to receive column_1 through column_4.",
        last_column=6,
    )
    raw_rows = [
        ("R-001", date(2026, 4, 2), 10, 5.25),
        ("R-002", date(2026, 4, 3), 15, 7.5),
        ("R-003", date(2026, 4, 7), 8, 12.0),
        ("R-004", date(2026, 4, 9), 20, 3.75),
    ]
    for row_index, row in enumerate(raw_rows, start=5):
        for column, value in enumerate(row, start=3):
            raw.cell(row_index, column, value)
        raw.cell(row_index, 4).number_format = "yyyy-mm-dd"
        raw.cell(row_index, 5).number_format = "#,##0"
        raw.cell(row_index, 6).number_format = '"$"#,##0.00'
        raw.row_dimensions[row_index].height = 22
    for cell in raw["C5:F8"][0]:
        cell.border = Border(top=Side(style="medium", color=TEAL))
    _style_body(raw, "C5:F8")
    for row in raw["C5:F8"]:
        for cell in row:
            if cell.row % 2 == 1:
                cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_TEAL)
    _set_widths(raw, {"A": 4, "B": 4, "C": 16, "D": 16, "E": 12, "F": 14})

    reference = f"{quote_sheetname(inventory.title)}!$B$4:$E$8"
    workbook.defined_names.add(DefinedName("InventoryData", attr_text=reference))

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "named_and_headerless.xlsx"
    workbook.save(path)
    workbook.close()
    return path


def build_legacy_scattered(output_dir: Path) -> Path:
    """Create an Excel 97-2003 workbook with scattered header columns."""

    workbook = LegacyWorkbook()
    sheet = workbook.add_sheet("Legacy Orders")
    title_style = easyxf(
        "font: bold on, colour white, height 320; "
        "pattern: pattern solid, fore_colour dark_blue; "
        "align: vert centre"
    )
    header_style = easyxf(
        "font: bold on, colour white; pattern: pattern solid, fore_colour teal; align: vert centre"
    )
    date_style = easyxf(num_format_str="YYYY-MM-DD")
    currency_style = easyxf(num_format_str='"$"#,##0.00')

    sheet.write_merge(0, 0, 0, 6, "Orders — Legacy XLS", title_style)
    sheet.write(1, 0, "Excel 97-2003 BIFF8 example with non-adjacent table columns.")
    headers = ["Customer ID", "Region", "Notes", "Invoice Date", "Owner", "Status", "Amount"]
    for column, value in enumerate(headers):
        sheet.write(3, column, value, header_style)

    rows = {
        4: ("L-001", "North", date(2026, 5, 1), "Ava", 1250.0),
        5: ("L-002", "South", date(2026, 5, 4), "Ben", 840.0),
        7: ("L-003", "West", date(2026, 5, 9), "Chloe", 2190.0),
    }
    for row, values in rows.items():
        customer_id, region, invoice_date, owner, amount = values
        sheet.write(row, 0, customer_id)
        sheet.write(row, 1, region)
        sheet.write(row, 3, invoice_date, date_style)
        sheet.write(row, 4, owner)
        sheet.write(row, 6, amount, currency_style)

    widths = {0: 16, 1: 14, 2: 10, 3: 16, 4: 16, 5: 12, 6: 15}
    for column, width in widths.items():
        sheet.col(column).width = width * 256
    sheet.col(3).hidden = True

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "legacy_scattered.xls"
    workbook.save(str(path))
    return path


def build_all(output_dir: Path = WORKBOOK_DIR) -> tuple[Path, ...]:
    return (
        build_native_table(output_dir),
        build_scattered_headers(output_dir),
        build_named_and_headerless(output_dir),
        build_legacy_scattered(output_dir),
    )


def main() -> None:
    for path in build_all():
        print(path.relative_to(Path.cwd()) if path.is_relative_to(Path.cwd()) else path)


if __name__ == "__main__":
    main()
