from __future__ import annotations

from pathlib import Path

import pytest
import xlwt
from openpyxl import Workbook

from excel_data_reader import (
    BodyPolicy,
    DiagnosticCode,
    ExcelDataReaderError,
    ExcelReader,
    MatrixBoundarySource,
    MatrixQuery,
    Severity,
    WorkbookFormat,
)


def _write_matrix_headers(sheet, *, leaf_row: int = 4) -> None:
    """Write the shared two-level grouped headers used by matrix tests.

    Args:
        sheet: OpenPyXL worksheet receiving the headers.
        leaf_row: One-based row receiving leaf attribute labels.
    """

    sheet.merge_cells(start_row=1, start_column=5, end_row=leaf_row - 1, end_column=7)
    sheet.merge_cells(start_row=1, start_column=8, end_row=leaf_row - 1, end_column=10)
    sheet.merge_cells(start_row=1, start_column=11, end_row=leaf_row - 1, end_column=12)
    sheet["E1"] = "group1"
    sheet["H1"] = "group2"
    sheet["K1"] = "group3"
    for column, value in enumerate(
        ("attr1", "attr2", "attr3", "attr1", "attr2", "attr3", "attr1", "attr2"),
        start=5,
    ):
        sheet.cell(leaf_row, column, value)


def _save_merged_matrix(path: Path) -> None:
    """Create a matrix workbook with vertically merged section anchors.

    Args:
        path: Destination path for the temporary workbook.
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matrix"
    _write_matrix_headers(sheet)
    sheet.merge_cells("B5:C7")
    sheet["B5"] = "Country Identifier"
    sheet.merge_cells("B10:C11")
    sheet["B10"] = "Sector Identifier"
    for row, identifier in enumerate(("Africa", "Philippines", "Korea"), start=5):
        sheet.cell(row, 4, identifier)
        for column in range(5, 13):
            sheet.cell(row, column, row * 100 + column)
    for row, identifier in enumerate(("Energy", "Health"), start=10):
        sheet.cell(row, 4, identifier)
        for column in range(5, 13):
            sheet.cell(row, column, row * 100 + column)
    workbook.save(path)
    workbook.close()


def test_merged_sections_resolve_hierarchical_headers_and_record_shapes(tmp_path: Path) -> None:
    path = tmp_path / "merged-matrix.xlsx"
    _save_merged_matrix(path)

    with ExcelReader.open(path) as reader:
        matches = reader.find_matrices(
            MatrixQuery(("Country Identifier", "Sector Identifier"), sheet="Matrix")
        )
        country = matches.require_section("country_identifier")
        sector = matches.require_section("sector identifier")
        data = reader.extract_matrix(country)

    assert not matches.diagnostics
    assert country.identifier_column == 4
    assert country.header_rows == (1, 4)
    assert country.boundary_source is MatrixBoundarySource.MERGED_SECTION
    assert country.range == "D5:L7"
    assert sector.range == "D10:L11"
    assert [header.labels for header in country.headers] == [
        ("group1", "attr1"),
        ("group1", "attr2"),
        ("group1", "attr3"),
        ("group2", "attr1"),
        ("group2", "attr2"),
        ("group2", "attr3"),
        ("group3", "attr1"),
        ("group3", "attr2"),
    ]
    assert len(data.values) == 24
    assert dict(data.long_records()[0]) == {
        "section": "Country Identifier",
        "identifier": "Africa",
        "group": "group1",
        "attribute": "attr1",
        "value": 505,
        "source_sheet": "Matrix",
        "source_cell": "E5",
        "source_row": 5,
        "source_column": 5,
        "identifier_cell": "D5",
    }
    assert dict(data.wide_records()[0]) == {
        "section": "Country Identifier",
        "identifier": "Africa",
        "group1__attr1": 505,
        "group1__attr2": 506,
        "group1__attr3": 507,
        "group2__attr1": 508,
        "group2__attr2": 509,
        "group2__attr3": 510,
        "group3__attr1": 511,
        "group3__attr2": 512,
        "source_sheet": "Matrix",
        "source_row": 5,
        "identifier_cell": "D5",
    }


def test_unmerged_sections_propagate_until_next_anchor_or_blank_rows(tmp_path: Path) -> None:
    path = tmp_path / "unmerged-sections.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matrix"
    sheet["E1"] = "group1"
    sheet["G1"] = "group2"
    sheet.append([None, None, None, None, "attr1", "attr2", "attr1", "attr2"])
    sheet["B3"] = "Country Identifier"
    sheet["D3"] = "Country1"
    sheet["D4"] = "Country2"
    sheet["B6"] = "Sector Identifier"
    sheet["D6"] = "Sector1"
    sheet["D7"] = "Sector2"
    for row in (3, 4, 6, 7):
        for column in range(5, 9):
            sheet.cell(row, column, row * 10 + column)
    sheet["A10"] = "unrelated note after two blank matrix rows"
    workbook.save(path)
    workbook.close()

    with ExcelReader.open(path) as reader:
        matches = reader.find_matrices(
            MatrixQuery(("Country Identifier", "Sector Identifier"), sheet="Matrix")
        )

    country = matches.require_section("Country Identifier")
    sector = matches.require_section("Sector Identifier")
    assert country.anchor.a1 == "B3"
    assert country.identifier_column == 4
    assert country.header_rows == (1, 2)
    assert country.range == "D3:H4"
    assert country.boundary_source is MatrixBoundarySource.NEXT_SECTION
    assert sector.range == "D6:H7"
    assert sector.boundary_source is MatrixBoundarySource.BLANK_ROWS
    assert [header.labels for header in country.headers] == [
        ("group1", "attr1"),
        ("group1", "attr2"),
        ("group2", "attr1"),
        ("group2", "attr2"),
    ]


def test_identifier_override_resolves_density_tie(tmp_path: Path) -> None:
    path = tmp_path / "identifier-tie.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matrix"
    sheet["E1"] = "group1"
    sheet["E2"] = "attr1"
    sheet["B3"] = "Country Identifier"
    sheet["C3"] = "C-1"
    sheet["D3"] = "Country1"
    sheet["C4"] = "C-2"
    sheet["D4"] = "Country2"
    sheet["E3"] = 1
    sheet["E4"] = 2
    workbook.save(path)
    workbook.close()

    with ExcelReader.open(path) as reader:
        ambiguous = reader.find_matrices(MatrixQuery(("Country Identifier",), sheet="Matrix"))
        resolved = reader.find_matrices(
            MatrixQuery(
                ("Country Identifier",),
                sheet="Matrix",
                identifier_column="D",
            )
        )

    assert not ambiguous.matches
    assert ambiguous.diagnostics[0].code is DiagnosticCode.AMBIGUOUS_IDENTIFIER_COLUMN
    with pytest.raises(ExcelDataReaderError) as captured:
        ambiguous.require_section("Country Identifier")
    assert captured.value.diagnostics[0].code is DiagnosticCode.AMBIGUOUS_IDENTIFIER_COLUMN
    assert resolved.require_section("Country Identifier").identifier_column == 4


def test_populated_row_without_identifier_is_warned_and_not_propagated(tmp_path: Path) -> None:
    path = tmp_path / "missing-identifier.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matrix"
    sheet["E1"] = "group1"
    sheet["E2"] = "attr1"
    sheet["B3"] = "Country Identifier"
    sheet["D3"] = "Country1"
    sheet["E3"] = 1
    sheet["E4"] = 2
    workbook.save(path)
    workbook.close()

    with ExcelReader.open(path) as reader:
        matches = reader.find_matrices(
            MatrixQuery(
                ("Country Identifier",),
                sheet="Matrix",
                identifier_column="D",
                body=BodyPolicy.last_populated(),
            )
        )
        match = matches.require_section("Country Identifier")
        data = reader.extract_matrix(match)

    assert match.diagnostics[0].code is DiagnosticCode.MISSING_ROW_IDENTIFIER
    assert match.diagnostics[0].severity is Severity.WARNING
    assert data.values[1].identifier is None


def test_section_aliases_within_and_repeated_anchors_remain_explicit(tmp_path: Path) -> None:
    path = tmp_path / "repeated-sections.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matrix"
    sheet["E1"] = "group1"
    sheet["E2"] = "attr1"
    sheet["B3"] = "Nation Identifier"
    sheet["D3"] = "Country1"
    sheet["E3"] = 1
    sheet["B6"] = "Nation Identifier"
    sheet["D6"] = "Country2"
    sheet["E6"] = 2
    workbook.save(path)
    workbook.close()

    query = MatrixQuery(
        ("Country Identifier",),
        aliases={"Country Identifier": ("Nation Identifier",)},
        sheet="Matrix",
    )
    with ExcelReader.open(path) as reader:
        ambiguous = reader.find_matrices(query)
        scoped = reader.find_matrices(
            MatrixQuery(
                ("Country Identifier",),
                aliases={"Country Identifier": ("Nation Identifier",)},
                sheet="Matrix",
                within="A1:E4",
            )
        )

    assert len(ambiguous.matches) == 2
    assert ambiguous.diagnostics[-1].code is DiagnosticCode.AMBIGUOUS_MATRIX_SECTION
    with pytest.raises(ExcelDataReaderError, match="matched 2 anchors"):
        ambiguous.require_section("Country Identifier")
    match = scoped.require_section("country_identifier")
    assert match.raw_section == "Nation Identifier"
    assert match.range == "D3:E3"


def test_duplicate_complete_header_paths_are_lossless_only_in_long_output(tmp_path: Path) -> None:
    path = tmp_path / "duplicate-paths.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matrix"
    sheet["E1"] = "group1"
    sheet["E2"] = "attr1"
    sheet["F2"] = "attr1"
    sheet["B3"] = "Country Identifier"
    sheet["D3"] = "Country1"
    sheet["E3"] = 1
    sheet["F3"] = 2
    workbook.save(path)
    workbook.close()

    with ExcelReader.open(path) as reader:
        match = reader.find_matrices(
            MatrixQuery(("Country Identifier",), sheet="Matrix")
        ).require_section("Country Identifier")
        data = reader.extract_matrix(match)

    assert [record["value"] for record in data.long_records()] == [1, 2]
    with pytest.raises(ExcelDataReaderError) as captured:
        data.wide_records()
    assert captured.value.diagnostics[0].code is DiagnosticCode.DUPLICATE_MATRIX_HEADER


def test_identifier_column_can_have_a_leaf_label_without_a_parent_group(tmp_path: Path) -> None:
    path = tmp_path / "labeled-identifier.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matrix"
    sheet.merge_cells("E1:F1")
    sheet["E1"] = "group1"
    sheet["D2"] = "Country"
    sheet["E2"] = "attr1"
    sheet["F2"] = "attr2"
    sheet["B3"] = "Country Identifier"
    sheet["D3"] = "Country1"
    sheet["E3"] = 1
    sheet["F3"] = 2
    workbook.save(path)
    workbook.close()

    with ExcelReader.open(path) as reader:
        match = reader.find_matrices(
            MatrixQuery(("Country Identifier",), sheet="Matrix")
        ).require_section("Country Identifier")

    assert match.identifier_column == 4
    assert [header.source_column for header in match.headers] == [5, 6]


def test_matrix_query_validates_level_names_rows_columns_and_aliases(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="at least one non-empty matrix section"):
        MatrixQuery(())
    with pytest.raises(ValueError, match="header_rows must contain"):
        MatrixQuery(("Country",), header_rows=(1,))
    with pytest.raises(ValueError, match="identifier_column"):
        MatrixQuery(("Country",), identifier_column="4")

    path = tmp_path / "aliases.xlsx"
    workbook = Workbook()
    workbook.save(path)
    workbook.close()
    with ExcelReader.open(path) as reader, pytest.raises(ExcelDataReaderError) as captured:
        reader.find_matrices(MatrixQuery(("Country",), aliases={"Unknown": ("Nation",)}))
    assert captured.value.diagnostics[0].code is DiagnosticCode.INVALID_MATRIX_QUERY


def test_matrix_discovery_uses_the_legacy_xls_adapter(tmp_path: Path) -> None:
    path = tmp_path / "matrix.xls"
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Matrix")
    sheet.write_merge(0, 2, 4, 5, "group1")
    sheet.write(3, 4, "attr1")
    sheet.write(3, 5, "attr2")
    sheet.write_merge(4, 5, 1, 2, "Country Identifier")
    sheet.write(4, 3, "Country1")
    sheet.write(5, 3, "Country2")
    sheet.write(4, 4, 10)
    sheet.write(4, 5, 20)
    sheet.write(5, 4, 30)
    sheet.write(5, 5, 40)
    workbook.save(str(path))

    with ExcelReader.open(path) as reader:
        match = reader.find_matrices(
            MatrixQuery(("Country Identifier",), sheet="Matrix")
        ).require_section("Country Identifier")
        data = reader.extract_matrix(match)
        workbook_format = reader.workbook_format

    assert workbook_format is WorkbookFormat.LEGACY_XLS
    assert match.header_rows == (1, 4)
    assert match.range == "D5:F6"
    assert [record["value"] for record in data.long_records()] == [10, 20, 30, 40]
