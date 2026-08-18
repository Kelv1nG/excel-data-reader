from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from excel_data_reader import (
    DiagnosticCode,
    ExcelDataReaderError,
    ExcelReader,
    FormulaValue,
    MatchSource,
    ValueMode,
    normalize_header,
)


@pytest.fixture
def workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "reader-fixture.xlsx"
    workbook = Workbook()

    native = workbook.active
    native.title = "Native"
    native.append(["Name", "Units", "Amount"])
    native.append(["Alpha", 2, "=B2*10"])
    native.append(["Beta", 3, "=B3*10"])
    native.append(["Gamma", 4, "=B4*10"])
    native.append(["Total", 9, "=SUM(C2:C4)"])
    native.add_table(
        Table(
            displayName="SalesTable",
            ref="A1:C5",
            totalsRowCount=1,
            totalsRowShown=True,
        )
    )

    scattered = workbook.create_sheet("Scattered")
    scattered["A3"] = " Customer\nID "
    scattered["D3"] = "Invoice-Date"
    scattered["G3"] = "Amount"
    scattered["A4"] = "C-001"
    scattered["D4"] = date(2026, 1, 2)
    scattered["G4"] = 10
    scattered["B5"] = "unrelated note"
    scattered["A6"] = "C-002"
    scattered["D6"] = date(2026, 1, 3)
    scattered["G6"] = 20

    named = workbook.create_sheet("Named Data")
    named["B2"] = "Code"
    named["C2"] = "Qty"
    named["D2"] = "Price"
    named["B3"] = "X"
    named["C3"] = 2
    named["D3"] = 5.5
    named["B4"] = "Y"
    named["C4"] = 3
    named["D4"] = 7.0
    workbook.defined_names.add(DefinedName("InvoiceData", attr_text="'Named Data'!$B$2:$D$4"))
    workbook.defined_names.add(
        DefinedName(
            "MultiArea",
            attr_text="'Named Data'!$B$2:$B$4,'Named Data'!$D$2:$D$4",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "DynamicData",
            attr_text="OFFSET('Named Data'!$B$2,0,0,3,3)",
        )
    )
    workbook.defined_names.add(DefinedName("ConstantValue", attr_text="42"))

    raw = workbook.create_sheet("Raw")
    raw["C5"] = "r1c1"
    raw["D5"] = 10
    raw["E5"] = True
    raw["C6"] = "r2c1"
    raw["D6"] = 20
    raw["E6"] = False

    duplicate = workbook.create_sheet("Duplicate")
    duplicate.append(["Name", " name "])
    duplicate.append(["Alpha", "A"])

    repeated = workbook.create_sheet("Repeated")
    repeated.append(["ID", "Amount"])
    repeated.append(["A", 1])
    repeated["A5"] = "ID"
    repeated["B5"] = "Amount"
    repeated["A6"] = "B"
    repeated["B6"] = 2

    sparse = workbook.create_sheet("Sparse")
    sparse["B2"] = "alpha"
    sparse["D4"] = 9
    sparse["E5"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    sparse.row_dimensions[4].hidden = True
    sparse.column_dimensions["D"].hidden = True

    workbook.save(path)
    workbook.close()
    return path


def test_header_normalization_is_exact_but_forgiving_about_separators() -> None:
    assert normalize_header("  Invoice\u00a0-Date\n") == "invoice date"
    assert normalize_header("Customer_ID") == "customer id"
    assert normalize_header("gross amount") != normalize_header("net amount")


def test_inventory_lists_structure(workbook_path: Path) -> None:
    with ExcelReader.open(workbook_path) as reader:
        inventory = reader.inventory()

    assert inventory.sheets[0].name == "Native"
    assert inventory.native_tables[0].name == "SalesTable"
    assert inventory.native_tables[0].bounds.a1 == "A1:C5"
    assert inventory.native_tables[0].totals_row_count == 1
    named = {item.name: item for item in inventory.named_ranges}
    assert named["InvoiceData"].destinations[0].a1 == "'Named Data'!B2:D4"
    assert named["DynamicData"].is_dynamic is True
    assert named["DynamicData"].is_resolvable is False
    assert named["ConstantValue"].is_dynamic is False
    assert named["ConstantValue"].is_resolvable is False


def test_native_table_extraction_and_structural_header_precedence(
    workbook_path: Path,
) -> None:
    with ExcelReader.open(workbook_path) as reader:
        table = reader.get_table("salestable")
        match = reader.find_tables(["amount", "name"], sheet="Native").require_one()
        projected = reader.extract(match)

    assert table.match.source is MatchSource.NATIVE_TABLE
    assert table.values[0] == ("Alpha", 2, "=B2*10")
    assert match.source is MatchSource.NATIVE_TABLE
    assert [column.source_column for column in match.columns] == [3, 1]
    assert projected.values[1] == ("=B3*10", "Beta")


def test_find_scattered_headers_preserves_requested_order_and_internal_blank_row(
    workbook_path: Path,
) -> None:
    with ExcelReader.open(workbook_path) as reader:
        match = reader.find_tables(
            ["amount", "customer id", "invoice date"],
            sheet="Scattered",
            max_blank_rows=2,
        ).require_one()
        table = reader.extract(match)

    assert match.range == "A3:G6"
    assert match.source is MatchSource.HEADER
    assert [column.source_column for column in match.columns] == [7, 1, 4]
    assert table.values == (
        (10, "C-001", datetime(2026, 1, 2)),
        (None, None, None),
        (20, "C-002", datetime(2026, 1, 3)),
    )


def test_non_adjacent_columns_can_be_rejected(workbook_path: Path) -> None:
    with ExcelReader.open(workbook_path) as reader:
        matches = reader.find_tables(
            ["customer id", "amount"],
            sheet="Scattered",
            allow_non_adjacent_columns=False,
        )

    assert not matches.matches
    with pytest.raises(ExcelDataReaderError) as captured:
        matches.require_one()
    assert captured.value.diagnostics[-1].code is DiagnosticCode.TABLE_NOT_FOUND


def test_repeated_headers_are_ambiguous(workbook_path: Path) -> None:
    with ExcelReader.open(workbook_path) as reader:
        matches = reader.find_tables(["id", "amount"], sheet="Repeated")

    assert len(matches.matches) == 2
    with pytest.raises(ExcelDataReaderError) as captured:
        matches.require_one()
    assert captured.value.diagnostics[0].code is DiagnosticCode.AMBIGUOUS_TABLE


def test_explicit_headerless_range_has_synthetic_columns(workbook_path: Path) -> None:
    with ExcelReader.open(workbook_path) as reader:
        table = reader.read_range("Raw", "C5:E6", header=None)

    assert [column.name for column in table.columns] == ["column_1", "column_2", "column_3"]
    assert table.values == (("r1c1", 10, True), ("r2c1", 20, False))
    assert table.rows[0].cells[0].address == "C5"
    assert dict(table.records()[1]) == {"column_1": "r2c1", "column_2": 20, "column_3": False}


def test_named_range_and_multi_destination_ambiguity(workbook_path: Path) -> None:
    with ExcelReader.open(workbook_path) as reader:
        table = reader.get_named_range("InvoiceData")
        multi = reader.find_named_ranges("MultiArea", header=None)

    assert table.match.source is MatchSource.NAMED_RANGE
    assert table.values == (("X", 2, 5.5), ("Y", 3, 7))
    assert len(multi.matches) == 2
    with pytest.raises(ExcelDataReaderError) as captured:
        multi.require_one()
    assert captured.value.diagnostics[0].code is DiagnosticCode.AMBIGUOUS_TABLE


def test_dynamic_named_range_is_reported(workbook_path: Path) -> None:
    with ExcelReader.open(workbook_path) as reader:
        matches = reader.find_named_ranges("DynamicData")

    with pytest.raises(ExcelDataReaderError) as captured:
        matches.require_one()
    assert captured.value.diagnostics[0].code is DiagnosticCode.DYNAMIC_NAMED_RANGE_UNRESOLVED


def test_constant_named_value_is_not_misclassified_as_dynamic(workbook_path: Path) -> None:
    with ExcelReader.open(workbook_path) as reader:
        matches = reader.find_named_ranges("ConstantValue")

    with pytest.raises(ExcelDataReaderError) as captured:
        matches.require_one()
    assert captured.value.diagnostics[0].code is DiagnosticCode.NON_RECTANGULAR_NAMED_RANGE


def test_duplicate_headers_remain_lossless_but_records_are_rejected(
    workbook_path: Path,
) -> None:
    with ExcelReader.open(workbook_path) as reader:
        table = reader.read_range("Duplicate", "A1:B2")

    assert table.values == (("Alpha", "A"),)
    with pytest.raises(ExcelDataReaderError) as captured:
        table.records()
    assert captured.value.diagnostics[0].code is DiagnosticCode.DUPLICATE_HEADER


def test_sparse_sheet_read_and_matrix(workbook_path: Path) -> None:
    with ExcelReader.open(workbook_path) as reader:
        sheet = reader.read_sheet("Sparse")
        with_styles = reader.read_sheet("Sparse", include_styled_blanks=True)

    assert [cell.address for cell in sheet.cells] == ["B2", "D4"]
    assert sheet.bounds is not None and sheet.bounds.a1 == "B2:D4"
    assert sheet.to_matrix() == (
        ("alpha", None, None),
        (None, None, None),
        (None, None, 9),
    )
    assert sheet.cells[1].hidden_row is True
    assert sheet.cells[1].hidden_column is True
    assert [cell.address for cell in with_styles.cells] == ["B2", "D4", "E5"]


@pytest.mark.parametrize(
    ("mode", "expected"),
    [
        (ValueMode.FORMULA, "=B2*10"),
        (ValueMode.CACHED, None),
        (ValueMode.BOTH, FormulaValue("=B2*10", None)),
    ],
)
def test_formula_value_modes(workbook_path: Path, mode: ValueMode, expected: object) -> None:
    with ExcelReader.open(workbook_path, value_mode=mode) as reader:
        table = reader.get_table("SalesTable")

    assert table.rows[0].cells[2].value == expected
    assert table.rows[0].cells[2].formula == "=B2*10"


def test_scan_limit_prevents_large_apparent_sheet_iteration(tmp_path: Path) -> None:
    path = tmp_path / "large-dimension.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["Z100"] = "far away"
    workbook.save(path)
    workbook.close()

    with (
        ExcelReader.open(path, max_scan_cells=1_000) as reader,
        pytest.raises(ExcelDataReaderError) as captured,
    ):
        reader.read_sheet("Sheet")

    assert captured.value.diagnostics[0].code is DiagnosticCode.SCAN_LIMIT_EXCEEDED


def test_closed_reader_fails_explicitly(workbook_path: Path) -> None:
    reader = ExcelReader.open(workbook_path)
    reader.close()

    with pytest.raises(ExcelDataReaderError) as captured:
        _ = reader.sheet_names

    assert captured.value.diagnostics[0].code is DiagnosticCode.READER_CLOSED
