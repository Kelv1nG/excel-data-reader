from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from excel_data_reader import (
    BodyPolicy,
    DiagnosticCode,
    ExcelDataReaderError,
    ExcelReader,
    MatchSource,
    TableQuery,
)


@pytest.fixture
def messy_workbook_path(tmp_path: Path) -> Path:
    """Build a compact corpus of awkward but common workbook layouts."""

    path = tmp_path / "messy-workbooks.xlsx"
    workbook = Workbook()

    aliases = workbook.active
    aliases.title = "Aliases"
    aliases.append(["Client No", None, "Gross Value", None, None, "Invoice Dt"])
    aliases.append(["C-001", None, 12.5, None, None, "2026-08-01"])
    aliases.append(["C-002", None, 20.0, None, None, "2026-08-02"])

    repeated = workbook.create_sheet("Repeated")
    repeated["B3"] = "ID"
    repeated["E3"] = "Amount"
    repeated["B4"] = "UP-1"
    repeated["E4"] = 10
    repeated["B5"] = "UP-2"
    repeated["E5"] = 20
    repeated["B12"] = "ID"
    repeated["E12"] = "Amount"
    repeated["B13"] = "DOWN-1"
    repeated["E13"] = 30

    tied = workbook.create_sheet("Tied")
    tied.append([])
    tied.append(["ID", "Amount"])
    tied.append(["UP", 1])
    tied["A7"] = "ID"
    tied["B7"] = "Amount"
    tied["A8"] = "DOWN"
    tied["B8"] = 2

    boundaries = workbook.create_sheet("Boundaries")
    boundaries["B2"] = "ID"
    boundaries["E2"] = "Amount"
    boundaries["B3"] = "A"
    boundaries["E3"] = 1
    boundaries["B4"] = "B"
    boundaries["E4"] = 2
    boundaries["B7"] = "C"
    boundaries["E7"] = 3
    boundaries["G12"] = "unrelated footer"

    native = workbook.create_sheet("Native")
    native.append(["ID", "Amount", "Notes"])
    native.append(["N-1", 5, "authored"])
    native.append(["N-2", 8, None])
    native.add_table(Table(displayName="NativeOrders", ref="A1:C3"))

    huge = workbook.create_sheet("Huge")
    huge["A1"] = "ID"
    huge["B1"] = "Amount"
    huge["A2"] = "H-1"
    huge["B2"] = 99
    huge["Z100"] = "inflates the apparent worksheet dimension"

    workbook.save(path)
    workbook.close()
    return path


def test_aliases_and_optional_headers_create_a_stable_logical_projection(
    messy_workbook_path: Path,
) -> None:
    query = TableQuery(
        required_headers=("customer id", "amount"),
        optional_headers=("invoice date", "owner"),
        aliases={
            "customer id": ("client no", "account number"),
            "amount": ("gross value",),
            "invoice date": ("invoice dt",),
        },
        sheet="Aliases",
    )

    with ExcelReader.open(messy_workbook_path) as reader:
        match = reader.query_tables(query).require_one()
        table = reader.extract(match)

    assert [column.requested_header for column in match.columns] == [
        "customer id",
        "amount",
        "invoice date",
    ]
    assert [column.source_column for column in match.columns] == [1, 3, 6]
    assert table.values[0] == ("C-001", 12.5, "2026-08-01")


def test_find_tables_accepts_a_table_query(messy_workbook_path: Path) -> None:
    query = TableQuery(
        required_headers=("customer id", "amount"),
        aliases={"customer id": "client no", "amount": "gross value"},
        sheet="Aliases",
    )

    with ExcelReader.open(messy_workbook_path) as reader:
        match = reader.find_tables(query).require_one()

    assert match.range == "A1:C3"


def test_repeated_headers_remain_ambiguous_without_a_location_hint(
    messy_workbook_path: Path,
) -> None:
    query = TableQuery(("id", "amount"), sheet="Repeated")

    with ExcelReader.open(messy_workbook_path) as reader:
        matches = reader.query_tables(query)

    assert [match.header_row for match in matches.matches] == [3, 12]
    with pytest.raises(ExcelDataReaderError) as captured:
        matches.require_one()
    assert captured.value.diagnostics[0].code is DiagnosticCode.AMBIGUOUS_TABLE


def test_near_selects_the_closest_repeated_table(messy_workbook_path: Path) -> None:
    query = TableQuery(("id", "amount"), sheet="Repeated", near="A12")

    with ExcelReader.open(messy_workbook_path) as reader:
        match = reader.query_tables(query).require_one()

    assert match.header_row == 12
    assert match.range == "B12:E13"


def test_near_keeps_equal_distance_matches_ambiguous(messy_workbook_path: Path) -> None:
    query = TableQuery(("id", "amount"), sheet="Tied", near="A5")

    with ExcelReader.open(messy_workbook_path) as reader:
        matches = reader.query_tables(query)

    assert [match.header_row for match in matches.matches] == [2, 7]
    with pytest.raises(ExcelDataReaderError) as captured:
        matches.require_one()
    assert captured.value.diagnostics[0].code is DiagnosticCode.AMBIGUOUS_TABLE


def test_within_limits_discovery_and_body_inference(messy_workbook_path: Path) -> None:
    query = TableQuery(("id", "amount"), sheet="Repeated", within="A1:F8")

    with ExcelReader.open(messy_workbook_path) as reader:
        match = reader.query_tables(query).require_one()

    assert match.header_row == 3
    assert match.data_end_row == 5


def test_blank_row_policy_stops_before_later_islands(messy_workbook_path: Path) -> None:
    query = TableQuery(
        ("id", "amount"),
        sheet="Boundaries",
        body=BodyPolicy.until_blank_rows(2),
    )

    with ExcelReader.open(messy_workbook_path) as reader:
        match = reader.query_tables(query).require_one()

    assert match.data_end_row == 4
    assert match.range == "B2:E4"


def test_larger_blank_tolerance_can_bridge_a_gap(messy_workbook_path: Path) -> None:
    query = TableQuery(
        ("id", "amount"),
        sheet="Boundaries",
        body=BodyPolicy.until_blank_rows(3),
    )

    with ExcelReader.open(messy_workbook_path) as reader:
        match = reader.query_tables(query).require_one()

    assert match.data_end_row == 7


def test_last_populated_policy_ignores_unrelated_columns(messy_workbook_path: Path) -> None:
    query = TableQuery(
        ("id", "amount"),
        sheet="Boundaries",
        body=BodyPolicy.last_populated(),
    )

    with ExcelReader.open(messy_workbook_path) as reader:
        match = reader.query_tables(query).require_one()

    assert match.data_end_row == 7


def test_explicit_body_policy_preserves_requested_blank_rows(
    messy_workbook_path: Path,
) -> None:
    query = TableQuery(
        ("id", "amount"),
        sheet="Boundaries",
        body=BodyPolicy.through_row(8),
    )

    with ExcelReader.open(messy_workbook_path) as reader:
        match = reader.query_tables(query).require_one()
        table = reader.extract(match)

    assert match.data_end_row == 8
    assert len(table.rows) == 6
    assert table.rows[-1].values == (None, None)


def test_native_table_keeps_authored_boundary_and_optional_columns(
    messy_workbook_path: Path,
) -> None:
    query = TableQuery(
        ("id", "amount"),
        optional_headers=("notes",),
        sheet="Native",
        body=BodyPolicy.through_row(1),
    )

    with ExcelReader.open(messy_workbook_path) as reader:
        match = reader.query_tables(query).require_one()

    assert match.source is MatchSource.NATIVE_TABLE
    assert match.range == "A1:C3"
    assert match.data_end_row == 3
    assert [column.requested_header for column in match.columns] == [
        "id",
        "amount",
        "notes",
    ]


def test_within_requires_a_native_table_to_be_fully_contained(
    messy_workbook_path: Path,
) -> None:
    query = TableQuery(("id", "amount"), sheet="Native", within="A1:B5")

    with ExcelReader.open(messy_workbook_path) as reader:
        match = reader.query_tables(query).require_one()

    assert match.source is MatchSource.HEADER
    assert match.range == "A1:B3"


def test_within_can_make_a_large_apparent_sheet_safe_to_scan(
    messy_workbook_path: Path,
) -> None:
    query = TableQuery(("id", "amount"), sheet="Huge", within="A1:B5")

    with ExcelReader.open(messy_workbook_path, max_scan_cells=10) as reader:
        match = reader.query_tables(query).require_one()

    assert match.range == "A1:B2"


@pytest.mark.parametrize(
    "query",
    [
        TableQuery((), sheet="Aliases"),
        TableQuery(("id",), aliases={"missing": ("identifier",)}, sheet="Aliases"),
        TableQuery(
            ("id", "amount"),
            aliases={"id": ("value",), "amount": ("value",)},
            sheet="Aliases",
        ),
    ],
)
def test_invalid_query_schemas_are_rejected(
    messy_workbook_path: Path,
    query: TableQuery,
) -> None:
    with (
        ExcelReader.open(messy_workbook_path) as reader,
        pytest.raises(ExcelDataReaderError) as captured,
    ):
        reader.query_tables(query)

    assert captured.value.diagnostics[0].code is DiagnosticCode.INVALID_HEADER_QUERY


def test_query_rejects_invalid_location_and_conflicting_legacy_options(
    messy_workbook_path: Path,
) -> None:
    with ExcelReader.open(messy_workbook_path) as reader:
        with pytest.raises(ExcelDataReaderError) as invalid_near:
            reader.query_tables(TableQuery(("id",), near="A1:B2"))
        with pytest.raises(ExcelDataReaderError) as conflicting:
            reader.find_tables(TableQuery(("id",)), sheet="Aliases")

    assert invalid_near.value.diagnostics[0].code is DiagnosticCode.INVALID_RANGE
    assert conflicting.value.diagnostics[0].code is DiagnosticCode.INVALID_HEADER_QUERY


def test_explicit_bottom_must_lie_inside_within(messy_workbook_path: Path) -> None:
    query = TableQuery(
        ("id", "amount"),
        sheet="Boundaries",
        within="A1:F7",
        body=BodyPolicy.through_row(8),
    )

    with (
        ExcelReader.open(messy_workbook_path) as reader,
        pytest.raises(ExcelDataReaderError) as captured,
    ):
        reader.query_tables(query)

    assert captured.value.diagnostics[0].code is DiagnosticCode.INVALID_HEADER_QUERY
